import streamlit as st
import pandas as pd
import numpy as np
from datetime import date, datetime
import pyodbc
import plotly.express as px
from io import BytesIO
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# GLOBAL CONSTANTS
# -------------------------------------------------
EXCLUDED_VENDOR = "ABASTECEDORA Y SUMINISTROS ORTEGA/ISABEL VALDEZ JIMENEZ"

# ✅ Base window MUST match Power BI query exactly
PBI_START = date(2025, 11, 1)
PBI_END = date(2026, 1, 31)  # ✅ Power BI M code uses '20260131'

# -------------------------------------------------
# CONFIG STREAMLIT
# -------------------------------------------------
st.set_page_config(
    page_title="Transito Globlal 2.0",
    page_icon="🚗",
    layout="wide",
)

# ---------- Fancy global styles ----------
st.markdown(
    """
<style>
html, body, [class*="css"]  {
    font-family: "Segoe UI", system-ui, sans-serif;
}

/* ---------- Design tokens (light default) ---------- */
:root {
    --accent-1: #0ea5e9;
    --accent-2: #6366f1;
    --accent-3: #22c55e;

    --card-radius: 16px;
    --pill-radius: 999px;

    /* Light theme defaults */
    --card-bg: rgba(255,255,255,0.9);
    --card-border: rgba(15,23,42,0.08);
    --card-shadow: 0 8px 20px rgba(15,23,42,0.08);

    --tab-bg: rgba(148,163,184,0.08);
    --tab-border: rgba(148,163,184,0.35);
    --tab-fg: #111827;

    --tab-bg-active: linear-gradient(90deg,#22c55e,#06b6d4);
    --tab-fg-active: #ffffff;
}

/* Override tokens when OS / browser is in dark mode */
@media (prefers-color-scheme: dark) {
    :root {
        --card-bg: rgba(15,23,42,0.9);
        --card-border: rgba(148,163,184,0.55);
        --card-shadow: 0 10px 30px rgba(15,23,42,0.7);

        --tab-bg: rgba(15,23,42,0.9);
        --tab-border: rgba(148,163,184,0.55);
        --tab-fg: #e5e7eb;

        --tab-bg-active: linear-gradient(90deg,#22c55e,#06b6d4);
        --tab-fg-active: #0f172a;
    }
}

/* Container spacing */
.block-container {
    padding-top: 1.2rem;
    padding-bottom: 2rem;
}

/* Title */
h1 {
    font-weight: 800 !important;
}

/* ---------- Tabs ---------- */
.stTabs [role="tablist"] {
    gap: 6px;
}
.stTabs [role="tab"] {
    padding: 6px 14px;
    border-radius: var(--pill-radius);
    background-color: var(--tab-bg);
    color: var(--tab-fg);
    border: 1px solid var(--tab-border);
    font-weight: 500;
}
.stTabs [aria-selected="true"] {
    background: var(--tab-bg-active);
    color: var(--tab-fg-active) !important;
    border-color: transparent !important;
}

/* ---------- Metrics ---------- */
[data-testid="stMetric"] {
    background: var(--card-bg);
    border-radius: var(--card-radius);
    padding: 10px 14px;
    border: 1px solid var(--card-border);
    box-shadow: var(--card-shadow);
}
[data-testid="stMetricValue"] {
    font-size: 1.7rem;
    font-weight: 800;
}

/* ---------- Custom metric for En tránsito ---------- */
.metric-alert {
    background: linear-gradient(90deg,#f97316,#ef4444);
    border-radius: var(--card-radius);
    padding: 10px 14px;
    border: 1px solid rgba(248,113,113,0.6);
    box-shadow: var(--card-shadow);
    color: #f9fafb;
}
.metric-alert-label {
    font-size: 0.9rem;
    opacity: 0.9;
}
.metric-alert-value {
    font-size: 1.7rem;
    font-weight: 800;
}

/* ---------- Download buttons ---------- */
div[data-testid="stDownloadButton"] > button {
    border-radius: var(--pill-radius);
    background: linear-gradient(90deg,var(--accent-1),var(--accent-2));
    color: #f9fafb;
    border: none;
    padding: 0.4rem 1.3rem;
    font-weight: 600;
}
div[data-testid="stDownloadButton"] > button:hover {
    filter: brightness(1.06);
}

/* ---------- Plotly transparent background ---------- */
.js-plotly-plot .plotly .main-svg {
    background-color: rgba(0,0,0,0) !important;
}
</style>
""",
    unsafe_allow_html=True,
)

# -------------------------------------------------
# SESSION STATE (control refresh manually)
# -------------------------------------------------
if "base_data" not in st.session_state:
    st.session_state["base_data"] = None

# -------------------------------------------------
# SMALL HELPER: DF -> EXCEL BYTES (auto-fit + filters)
# -------------------------------------------------
def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Datos") -> bytes:
    """Return an .xlsx file (bytes) with autofilter and auto column width."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column

        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)
    return output.getvalue()

# -------------------------------------------------
# MULTI-SHEET EXCEL (Resumen download incl. En Tránsito detail)
# -------------------------------------------------
def _safe_sheet_name(name: str) -> str:
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    out = str(name)
    for ch in invalid:
        out = out.replace(ch, "_")
    out = out.strip() or "Sheet"
    return out[:31]

def dfs_to_excel_bytes(sheets: dict) -> bytes:
    """Return an .xlsx (bytes) with multiple sheets, each with autofilter + auto column width."""
    output = BytesIO()
    used_names = set()

    def unique_sheet_name(base: str) -> str:
        base = _safe_sheet_name(base)
        if base not in used_names:
            used_names.add(base)
            return base
        i = 2
        while True:
            suffix = f"_{i}"
            cand = (base[: (31 - len(suffix))] + suffix) if len(base) + len(suffix) > 31 else (base + suffix)
            if cand not in used_names:
                used_names.add(cand)
                return cand
            i += 1

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for raw_name, df in sheets.items():
            sheet_name = unique_sheet_name(raw_name)
            if df is None:
                df = pd.DataFrame()

            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            max_row = ws.max_row
            max_col = ws.max_column

            if max_col > 0 and max_row > 0:
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    for cell in ws[col_letter]:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

    output.seek(0)
    return output.getvalue()

# -------------------------------------------------
# ✅ HELPER (ONLY for Back Office tab): parse Back Office datetime robustly
# -------------------------------------------------
def parse_backoffice_datetime(series: pd.Series, window_start: date | None = None, window_end: date | None = None) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.replace({"nan": "", "None": "", "NaT": ""})
    s = s.where(s != "", np.nan)

    if s.notna().any():
        pat = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s+\d{1,2}:\d{2}(?::\d{2})?)|(\d{4}[/-]\d{1,2}[/-]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?)"
        ext = s.astype(str).str.extract(pat)
        ext = ext[0].fillna(ext[1])
        s2 = ext.where(ext.notna(), s)
    else:
        s2 = s

    dt_dayfirst = pd.to_datetime(s2, errors="coerce", dayfirst=True)
    dt_monthfirst = pd.to_datetime(s2, errors="coerce", dayfirst=False)

    if window_start is None or window_end is None:
        return dt_dayfirst

    w0 = pd.Timestamp(window_start)
    w1 = pd.Timestamp(window_end) + pd.Timedelta(days=1) - pd.Timedelta(microseconds=1)

    in1 = dt_dayfirst.between(w0, w1)
    in2 = dt_monthfirst.between(w0, w1)

    out = dt_dayfirst.copy()
    out = out.where(~(in2 & ~in1), dt_monthfirst)
    out = out.where(~(dt_dayfirst.isna() & dt_monthfirst.notna()), dt_monthfirst)

    return out

# -------------------------------------------------
# DB CONNECTION
# -------------------------------------------------
@st.cache_resource
def get_connection():
    cfg = st.secrets["sql"]
    driver = cfg["driver"]
    server = cfg["server"]
    database = cfg["database"]
    user = cfg["user"]
    password = cfg["password"]

    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={user};"
        f"PWD={password};"
        "Encrypt=yes;"
        "TrustServerCertificate=yes;"
        "MARS_Connection=yes;"
    )
    return pyodbc.connect(conn_str, autocommit=True)

# -------------------------------------------------
# LOAD DATA FROM SQL
# -------------------------------------------------
@st.cache_data
def load_hoja1():
    sql = """
    SELECT DISTINCT
        e.[Nombre Completo] AS NombreCompleto,
        e.[Jefe Inmediato]  AS JefeDirecto,
        e.[Region],
        e.[SubRegion],
        e.[Plaza],
        e.[Tienda],
        e.[Puesto],
        e.[Canal de Venta],
        e.[Tipo Tienda],
        e.[Operacion],
        e.[Estatus]
    FROM reporte_empleado('EMPRESA_MAESTRA',1,'','') AS e
    WHERE
        e.[Canal de Venta] = 'ATT'
        AND e.[Operacion]   = 'CONTACT CENTER'
        AND e.[Tipo Tienda] = 'VIRTUAL'
        AND e.[Puesto] IN (
            'ASESOR TELEFONICO',
            'ASESOR TELEFONICO 7500',
            'EJECUTIVO TELEFONICO 6500 AM',
            'EJECUTIVO TELEFONICO 6500 PM',
            'SUPERVISOR DE CONTACT CENTER'
        )
        AND e.[Estatus] = 'ACTIVO';
    """
    conn = get_connection()
    df = pd.read_sql(sql, conn)

    text_cols = [
        "NombreCompleto","JefeDirecto","Region","SubRegion","Plaza","Tienda",
        "Puesto","Canal de Venta","Tipo Tienda","Operacion","Estatus"
    ]
    for col in text_cols:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": np.nan, "None": np.nan})

    df["JefeDirecto"] = df["JefeDirecto"].fillna("").str.strip()
    df["JefeDirecto"] = df["JefeDirecto"].replace("", "ENCUBADORA")
    df["Coordinador"] = df["JefeDirecto"]

    df = df[df["NombreCompleto"].str.upper() != EXCLUDED_VENDOR].copy()
    return df

@st.cache_data
def load_consulta1(fecha_ini: date, fecha_fin: date) -> pd.DataFrame:
    fi = fecha_ini.strftime("%Y%m%d")
    ff = fecha_fin.strftime("%Y%m%d")

    sql = f"""
    SELECT
        *,
        [Tienda solicita] AS Centro
    FROM reporte_programacion_entrega('empresa_maestra', 4, '{fi}', '{ff}')
    WHERE
        [Tienda solicita] LIKE 'EXP ATT C CENTER%' AND
        [Estatus] IN ('En entrega','Canc Error','Entregado',
                      'En preparacion','Back Office','Solicitado');
    """
    conn = get_connection()
    df = pd.read_sql(sql, conn)
    return df

# -------------------------------------------------
# TRANSFORMACIONES
# -------------------------------------------------
def transform_consulta1(df_raw: pd.DataFrame, hoja: pd.DataFrame) -> pd.DataFrame:
    df = df_raw.copy()

    for col in ["Centro", "Estatus", "Back Office", "Vendedor", "Cliente"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": np.nan, "None": np.nan})

    if "Venta" in df.columns:
        df["Venta"] = df["Venta"].replace({"nan": np.nan, "None": np.nan})

    if "Vendedor" in df.columns:
        df = df[df["Vendedor"].str.upper() != EXCLUDED_VENDOR].copy()

    df["Centro Original"] = np.nan
    mask_cc2 = df["Centro"].str.contains("EXP ATT C CENTER 2", na=False)
    mask_juarez = df["Centro"].str.contains("EXP ATT C CENTER JUAREZ", na=False)
    df.loc[mask_cc2, "Centro Original"] = "CC2"
    df.loc[mask_juarez, "Centro Original"] = "CC JV"

    def region_from_centro(c):
        if not isinstance(c, str):
            return np.nan
        if "GDL" in c:
            return "GDL"
        if "MEX" in c:
            return "MEX"
        if "MTY" in c:
            return "MTY"
        if "PUE" in c:
            return "PUE"
        if "TIJ" in c:
            return "TIJ"
        if "VER" in c:
            return "VER"
        return np.nan

    df["Region"] = df["Centro"].apply(region_from_centro)

    empleados_join = hoja[
        hoja["Puesto"].isin(["ASESOR TELEFONICO 7500", "EJECUTIVO TELEFONICO 6500 AM"])
    ].copy()
    empleados_join = empleados_join[empleados_join["JefeDirecto"] != "ENCUBADORA"]
    empleados_join = empleados_join[empleados_join["NombreCompleto"].str.upper() != EXCLUDED_VENDOR]
    empleados_join = empleados_join.drop_duplicates(subset=["NombreCompleto"])

    hoja_join = empleados_join.rename(
        columns={
            "NombreCompleto": "Nombre Completo",
            "JefeDirecto": "Jefe directo",
        }
    )

    df = df.merge(
        hoja_join[["Nombre Completo", "Jefe directo", "Coordinador"]],
        how="left",
        left_on="Vendedor",
        right_on="Nombre Completo",
    )
    df.drop(columns=["Nombre Completo"], inplace=True, errors="ignore")

    def status_calc(row):
        est = row.get("Estatus")
        venta = row.get("Venta")

        if est in ("En entrega", "En preparacion", "Solicitado", "Back Office"):
            return "En Transito"

        venta_vacia = pd.isna(venta) or (isinstance(venta, str) and venta == "")

        if est == "Entregado" and venta_vacia:
            return "En Transito"

        return "Entregado"

    df["Status"] = df.apply(status_calc, axis=1)

    df["Fecha creacion"] = pd.to_datetime(df["Fecha creacion"], errors="coerce", dayfirst=True)
    df["Fecha"] = df["Fecha creacion"].dt.date
    df["Hora"] = df["Fecha creacion"].dt.hour

    iso = df["Fecha creacion"].dt.isocalendar()
    df["Año"] = df["Fecha creacion"].dt.year
    df["MesNum"] = df["Fecha creacion"].dt.month
    df["Mes"] = df["Fecha creacion"].dt.strftime("%B")
    df["AñoMes"] = df["Fecha creacion"].dt.strftime("%Y-%m")
    df["Día"] = df["Fecha creacion"].dt.day
    df["Nombre Día"] = df["Fecha creacion"].dt.strftime("%A")

    df["Año Semana"] = iso["year"].astype(str) + "-W" + iso["week"].astype(str).str.zfill(2)

    df["Fecha contacto"] = pd.to_datetime(df["Fecha contacto"], errors="coerce", dayfirst=True)
    df["MesContactoNum"] = df["Fecha contacto"].dt.month

    df["Jefe directo"] = df["Jefe directo"].fillna("").astype(str).str.strip()
    df["Jefe directo"] = df["Jefe directo"].replace("", "ENCUBADORA")

    return df

# -------------------------------------------------
# SIN VENTA
# -------------------------------------------------
@st.cache_data
def build_sin_venta(hoja: pd.DataFrame, consulta: pd.DataFrame, ref_date: date) -> pd.DataFrame:
    empleados_sinv = hoja[
        hoja["Puesto"].isin(["ASESOR TELEFONICO 7500", "EJECUTIVO TELEFONICO 6500 AM"])
    ].copy()

    empleados_sinv = empleados_sinv[empleados_sinv["JefeDirecto"] != "ENCUBADORA"]
    empleados_sinv = empleados_sinv[empleados_sinv["NombreCompleto"].str.upper() != EXCLUDED_VENDOR]
    empleados_sinv = empleados_sinv.drop_duplicates(subset=["NombreCompleto"])

    year_ref = ref_date.year
    month_ref = ref_date.month
    fechas = pd.to_datetime(consulta["Fecha creacion"], errors="coerce")
    mask_mes = (fechas.dt.year == year_ref) & (fechas.dt.month == month_ref)
    ventas_mes = consulta.loc[mask_mes].copy()

    valid_status = ["Back Office","En entrega","En preparacion","Entregado","Solicitado"]
    ventas_validas = ventas_mes[ventas_mes["Estatus"].isin(valid_status)][["Vendedor"]].drop_duplicates()

    tmp = empleados_sinv.merge(
        ventas_validas,
        how="left",
        left_on="NombreCompleto",
        right_on="Vendedor",
        indicator=True,
    )
    sinv = tmp[tmp["_merge"] == "left_only"].copy()
    sinv.drop(columns=["Vendedor", "_merge"], inplace=True)

    return sinv

# -------------------------------------------------
# KPI HELPERS
# -------------------------------------------------
def kpi_activadas(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    return int((df["Status"] == "Entregado").sum())

def kpi_back(df: pd.DataFrame) -> int:
    return int((df["Estatus"] == "Back Office").sum())

def kpi_en_entrega(df: pd.DataFrame) -> int:
    return int((df["Estatus"] == "En entrega").sum())

def kpi_en_transito(df: pd.DataFrame) -> int:
    return int((df["Status"] == "En Transito").sum())

def kpi_preparacion(df: pd.DataFrame) -> int:
    return int((df["Estatus"] == "En preparacion").sum())

def kpi_solicitados(df: pd.DataFrame) -> int:
    return int((df["Estatus"] == "Solicitado").sum())

def kpi_total_sinventa(df_sinventa: pd.DataFrame) -> int:
    return int(df_sinventa.shape[0])

def kpi_validacion_pbi_all(ventasnc_all: pd.DataFrame) -> int:
    if ventasnc_all.empty or "Estatus" not in ventasnc_all.columns or "Venta" not in ventasnc_all.columns:
        return 0

    est_ok = ventasnc_all["Estatus"].astype(str).str.strip().eq("Entregado")
    venta = ventasnc_all["Venta"]
    blank_or_trim_empty = venta.isna() | venta.astype(str).str.strip().eq("")
    total = int((est_ok & blank_or_trim_empty).sum())
    return total

# -------------------------------------------------
# MAIN APP
# -------------------------------------------------
def main():
    st.title("Dashboard Transito Global  – CC")

    st.sidebar.header("Filtros")

    if st.sidebar.button("🔄 Actualizar datos"):
        st.cache_data.clear()
        st.cache_resource.clear()
        st.session_state["base_data"] = None
        st.rerun()

    default_start = PBI_START
    default_end = min(date.today(), PBI_END)

    fecha_ini = st.sidebar.date_input("Fecha inicio", default_start)
    fecha_fin = st.sidebar.date_input("Fecha fin", default_end)

    if fecha_ini > fecha_fin:
        st.sidebar.error("La fecha inicio no puede ser mayor que la fecha fin.")
        return

    if st.session_state["base_data"] is None:
        with st.spinner("Cargando datos desde SQL..."):
            hoja = load_hoja1()
            consulta_raw_base = load_consulta1(PBI_START, PBI_END)
            consulta_base = transform_consulta1(consulta_raw_base, hoja)
            validacion_pbi = kpi_validacion_pbi_all(consulta_base)

            st.session_state["base_data"] = {
                "hoja": hoja,
                "consulta_base": consulta_base,
                "validacion_pbi": validacion_pbi,
            }

    hoja = st.session_state["base_data"]["hoja"]
    consulta_base = st.session_state["base_data"]["consulta_base"]
    validacion_pbi = st.session_state["base_data"]["validacion_pbi"]

    consulta = consulta_base[
        (consulta_base["Fecha"] >= fecha_ini) & (consulta_base["Fecha"] <= fecha_fin)
    ].copy()

    sinventa = build_sin_venta(hoja, consulta, fecha_fin)

    centros = ["All"] + sorted([c for c in consulta["Centro Original"].dropna().unique().tolist()])
    supervisores = ["All"] + sorted([s for s in consulta["Jefe directo"].dropna().unique().tolist()])
    meses = ["All"] + sorted(consulta["Mes"].unique().tolist())

    centro_sel = st.sidebar.selectbox("Centro", centros, index=0)
    supervisor_sel = st.sidebar.selectbox("Supervisor", supervisores, index=0)
    mes_sel = st.sidebar.selectbox("Mes (Fecha creación)", meses, index=0)

    df_for_exec = consulta.copy()
    if centro_sel != "All":
        df_for_exec = df_for_exec[df_for_exec["Centro Original"] == centro_sel]
    if supervisor_sel != "All":
        df_for_exec = df_for_exec[df_for_exec["Jefe directo"] == supervisor_sel]
    if mes_sel != "All":
        df_for_exec = df_for_exec[df_for_exec["Mes"] == mes_sel]

    df_for_exec = df_for_exec[df_for_exec["Vendedor"].str.upper() != EXCLUDED_VENDOR]
    ejecutivos = ["All"] + sorted([e for e in df_for_exec["Vendedor"].dropna().unique().tolist()])
    ejecutivo_sel = st.sidebar.selectbox("Ejecutivo", ejecutivos, index=0)

    df_no_month = consulta.copy()
    if centro_sel != "All":
        df_no_month = df_no_month[df_no_month["Centro Original"] == centro_sel]
    if supervisor_sel != "All":
        df_no_month = df_no_month[df_no_month["Jefe directo"] == supervisor_sel]
    if ejecutivo_sel != "All":
        df_no_month = df_no_month[df_no_month["Vendedor"] == ejecutivo_sel]

    df = df_no_month.copy()
    if mes_sel != "All":
        df = df[df["Mes"] == mes_sel]

    sinv_fil = sinventa.copy()
    if supervisor_sel != "All":
        sinv_fil = sinv_fil[sinv_fil["JefeDirecto"] == supervisor_sel]
    sinv_fil = sinv_fil[sinv_fil["JefeDirecto"] != "ENCUBADORA"]

    tabs = st.tabs(
        [
            "Resumen",
            "Back Office",
            "Canceladas",
            "Programadas x semana",
            "Programadas (Top Ejecutivos)",
            "Detalle General",
            "Sin Venta",
        ]
    )

    # ==================== TAB 0: RESUMEN ====================
    with tabs[0]:
        st.subheader("Resumen de estatus")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("En preparación", kpi_preparacion(df))
            st.metric("Solicitados", kpi_solicitados(df))
        with col2:
            st.metric("En entrega", kpi_en_entrega(df))

            en_t = kpi_en_transito(df)
            st.markdown(
                f"""
                <div class="metric-alert">
                    <div class="metric-alert-label">En tránsito</div>
                    <div class="metric-alert-value">{en_t}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with col3:
            st.metric("Activadas (Entregado)", kpi_activadas(df))
            st.metric("Back Office", kpi_back(df))

        st.metric("Entregados sin venta (Validación)", validacion_pbi)

        cols_en_t_resumen = [
            c for c in [
                "Jefe directo",
                "Vendedor",
                "Cliente",
                "Telefono",
                "Folio",
                "Fecha",
                "Hora",
                "Centro",
                "Estatus",
                "Status",
                "Back Office",
                "Venta",
            ]
            if c in df.columns
        ]
        df_en_t_resumen = df[df["Status"] == "En Transito"][cols_en_t_resumen].copy()
        df_en_t_resumen = df_en_t_resumen.rename(
            columns={
                "Vendedor": "Ejecutivo",
                "Telefono": "Telefono cliente",
            }
        )

        st.download_button(
            "Descargar detalle (Excel)",
            data=dfs_to_excel_bytes(
                {
                    "Detalle": df,
                    "EnTransitoDetalle": df_en_t_resumen,
                }
            ),
            file_name=f"detalle_programacion_{fecha_ini}_{fecha_fin}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ==================== TAB 1: BACK OFFICE ====================
    with tabs[1]:
        st.subheader("Back Office")

        if "Back Office" not in consulta_base.columns:
            st.info("No existe la columna 'Back Office' en los datos.")
        else:
            df_bo_ctx = consulta_base.copy()

            if centro_sel != "All":
                df_bo_ctx = df_bo_ctx[df_bo_ctx["Centro Original"] == centro_sel]
            if supervisor_sel != "All":
                df_bo_ctx = df_bo_ctx[df_bo_ctx["Jefe directo"] == supervisor_sel]
            if ejecutivo_sel != "All":
                df_bo_ctx = df_bo_ctx[df_bo_ctx["Vendedor"] == ejecutivo_sel]

            bo_dt = parse_backoffice_datetime(df_bo_ctx["Back Office"], window_start=fecha_ini, window_end=fecha_fin)

            df_back = df_bo_ctx[(df_bo_ctx["Estatus"] != "Canc Error") & (bo_dt.notna())].copy()
            df_back["BO_DT"] = bo_dt
            df_back["BO_Fecha"] = df_back["BO_DT"].dt.date
            df_back["BO_Hora"] = df_back["BO_DT"].dt.hour

            df_back = df_back[(df_back["BO_Fecha"] >= fecha_ini) & (df_back["BO_Fecha"] <= fecha_fin)].copy()

            if mes_sel != "All" and not df_back.empty:
                df_back = df_back[df_back["BO_DT"].dt.strftime("%B") == mes_sel].copy()

            if df_back.empty:
                st.info("No hay registros Back Office (por fecha/hora de Rastreo) dentro del rango seleccionado.")
            else:
                by_day = df_back.groupby("BO_Fecha", as_index=False).size()
                fig = px.bar(
                    by_day,
                    x="BO_Fecha",
                    y="size",
                    title="Total por día (Back Office) — por fecha/hora de Back Office (Rastreo)",
                    labels={"size": "Total Back Office", "BO_Fecha": "Fecha Back Office"},
                )
                st.plotly_chart(fig, width="stretch")

                df_back["BO_MonthKey"] = df_back["BO_DT"].dt.strftime("%Y-%m")
                df_back["BO_MonthName"] = df_back["BO_DT"].dt.strftime("%B")
                df_back["BO_MonthLabel"] = df_back["BO_MonthKey"] + " (" + df_back["BO_MonthName"] + ")"

                month_start = df_back["BO_DT"].dt.to_period("M").dt.to_timestamp()
                first_wd = month_start.dt.weekday
                df_back["BO_WeekOfMonth"] = ((df_back["BO_DT"].dt.day + first_wd - 1) // 7) + 1

                st.markdown("### Vista por meses y semanas (Back Office)")

                month_options = sorted(df_back["BO_MonthLabel"].dropna().unique().tolist())
                default_months = month_options if month_options else []
                months_sel = st.multiselect(
                    "Selecciona uno o más meses (Back Office)",
                    options=month_options,
                    default=default_months,
                    key="bo_months_multi",
                )

                df_mw = df_back.copy()
                if months_sel:
                    df_mw = df_mw[df_mw["BO_MonthLabel"].isin(months_sel)].copy()
                else:
                    df_mw = df_mw.iloc[0:0].copy()

                if df_mw.empty:
                    st.info("No hay datos Back Office para los meses seleccionados.")
                else:
                    df_mw["BO_WeekLabel"] = df_mw["BO_MonthLabel"] + " - Semana " + df_mw["BO_WeekOfMonth"].astype(int).astype(str)
                    week_options = sorted(df_mw["BO_WeekLabel"].dropna().unique().tolist())
                    default_weeks = week_options if week_options else []

                    weeks_sel = st.multiselect(
                        "Selecciona Semana(s) del mes (Back Office)",
                        options=week_options,
                        default=default_weeks,
                        key="bo_weeks_multi",
                    )

                    if weeks_sel:
                        df_mw = df_mw[df_mw["BO_WeekLabel"].isin(weeks_sel)].copy()
                    else:
                        df_mw = df_mw.iloc[0:0].copy()

                    if df_mw.empty:
                        st.info("No hay datos Back Office para las semanas seleccionadas.")
                    else:
                        by_day_mw = df_mw.groupby("BO_Fecha", as_index=False).size()
                        fig_mw = px.bar(
                            by_day_mw,
                            x="BO_Fecha",
                            y="size",
                            title="Total por día (Back Office) — filtro por Mes(es) y Semana(s)",
                            labels={"size": "Total Back Office", "BO_Fecha": "Fecha Back Office"},
                        )
                        st.plotly_chart(fig_mw, width="stretch")

                        st.markdown("### Comparativo día vs día (mes contra mes)")

                        df_cmp = df_mw.copy()
                        df_cmp["BO_DiaDelMes"] = df_cmp["BO_DT"].dt.day

                        cmp = (
                            df_cmp.groupby(["BO_MonthLabel", "BO_DiaDelMes"], as_index=False)
                            .size()
                            .rename(columns={"size": "Total"})
                        )

                        fig_cmp = px.line(
                            cmp,
                            x="BO_DiaDelMes",
                            y="Total",
                            color="BO_MonthLabel",
                            markers=True,
                            title="Comparativo por día del mes (Back Office)",
                            labels={
                                "BO_DiaDelMes": "Día del mes",
                                "Total": "Total Back Office",
                                "BO_MonthLabel": "Mes",
                            },
                        )
                        fig_cmp.update_xaxes(dtick=1)
                        st.plotly_chart(fig_cmp, width="stretch")

                        st.markdown("#### Comparar dos fechas específicas (calendario)")

                        avail_dates = sorted(df_mw["BO_Fecha"].dropna().unique().tolist())
                        if not avail_dates:
                            st.info("No hay fechas disponibles para comparar con los filtros actuales.")
                        else:
                            d_min = avail_dates[0]
                            d_max = avail_dates[-1]
                            d_def_2 = d_max
                            d_def_1 = avail_dates[-2] if len(avail_dates) >= 2 else d_max

                            cA, cB = st.columns(2)
                            with cA:
                                d1 = st.date_input(
                                    "Fecha 1 (día a comparar)",
                                    value=d_def_1,
                                    min_value=d_min,
                                    max_value=d_max,
                                    key="bo_cmp_calendar_date1",
                                )
                            with cB:
                                d2 = st.date_input(
                                    "Fecha 2 (día a comparar)",
                                    value=d_def_2,
                                    min_value=d_min,
                                    max_value=d_max,
                                    key="bo_cmp_calendar_date2",
                                )

                            df_d1 = df_mw[df_mw["BO_Fecha"] == d1].copy()
                            df_d2 = df_mw[df_mw["BO_Fecha"] == d2].copy()

                            t1 = int(df_d1.shape[0])
                            t2 = int(df_d2.shape[0])

                            m1, m2, m3 = st.columns(3)
                            with m1:
                                st.metric("Total Back Office (Fecha 1)", t1)
                            with m2:
                                st.metric("Total Back Office (Fecha 2)", t2)
                            with m3:
                                st.metric("Diferencia (Fecha 1 - Fecha 2)", t1 - t2)

                            comp_df = pd.DataFrame(
                                {
                                    "Fecha": [str(d1), str(d2)],
                                    "Total": [t1, t2],
                                }
                            )
                            fig_dates = px.bar(
                                comp_df,
                                x="Fecha",
                                y="Total",
                                title="Comparativo Back Office — Fecha vs Fecha",
                                labels={"Total": "Total Back Office"},
                            )
                            st.plotly_chart(fig_dates, width="stretch")

                            h1 = df_d1.groupby("BO_Hora").size()
                            h2 = df_d2.groupby("BO_Hora").size()

                            hours = list(range(0, 24))
                            hour_df = pd.DataFrame(
                                {
                                    "Hora": hours,
                                    str(d1): [int(h1.get(h, 0)) for h in hours],
                                    str(d2): [int(h2.get(h, 0)) for h in hours],
                                }
                            )
                            hour_long = hour_df.melt(id_vars="Hora", var_name="Fecha", value_name="Total")

                            fig_hour = px.bar(
                                hour_long,
                                x="Hora",
                                y="Total",
                                color="Fecha",
                                barmode="group",
                                title="Comparativo por hora — Fecha vs Fecha (Back Office)",
                                labels={"Total": "Total Back Office", "Hora": "Hora Back Office"},
                            )
                            st.plotly_chart(fig_hour, width="stretch")

                day_options = sorted(by_day["BO_Fecha"].unique())
                today = date.today()
                default_index = day_options.index(today) if today in day_options else len(day_options) - 1

                day_sel = st.selectbox(
                    "Selecciona un día para ver el desglose por hora y equipo",
                    day_options,
                    index=default_index,
                    key="bo_day_sel",
                )

                df_day = df_back[df_back["BO_Fecha"] == day_sel].copy()

                by_hour_total = df_day.groupby("BO_Hora", as_index=False).size()
                fig_total = px.bar(
                    by_hour_total,
                    x="BO_Hora",
                    y="size",
                    title=f"Total Back Office por hora – {day_sel} (hora Back Office)",
                    labels={"size": "Total Back Office", "BO_Hora": "Hora Back Office"},
                )
                st.plotly_chart(fig_total, width="stretch")

                by_hour_team = (
                    df_day.groupby(["BO_Hora", "Jefe directo"], as_index=False)
                    .size()
                    .rename(columns={"size": "Total"})
                )

                fig_team = px.bar(
                    by_hour_team,
                    x="BO_Hora",
                    y="Total",
                    color="Jefe directo",
                    barmode="group",
                    title=f"Back Office por hora y equipo – {day_sel} (hora Back Office)",
                    labels={
                        "Total": "Total Back Office",
                        "BO_Hora": "Hora Back Office",
                        "Jefe directo": "Supervisor",
                    },
                )
                st.plotly_chart(fig_team, width="stretch")

                st.subheader("Detalle Back Office (Ejecutivo / Jefe directo)")
                detalle_cols_bo = [
                    c
                    for c in [
                        "Jefe directo",
                        "Vendedor",
                        "Cliente",
                        "Telefono",
                        "Folio",
                        "BO_Fecha",
                        "BO_Hora",
                        "Centro",
                        "Estatus",
                        "Back Office",
                        "Venta",
                    ]
                    if c in df_day.columns
                ]

                df_det_bo = df_day[detalle_cols_bo].rename(
                    columns={
                        "Vendedor": "Ejecutivo",
                        "Telefono": "Telefono cliente",
                        "BO_Fecha": "Fecha Back Office",
                        "BO_Hora": "Hora Back Office",
                    }
                )
                df_det_bo = df_det_bo.sort_values(
                    [col for col in ["Jefe directo", "Ejecutivo", "Fecha Back Office", "Hora Back Office", "Folio"] if col in df_det_bo.columns]
                )

                st.dataframe(df_det_bo, width="stretch")

                st.download_button(
                    "Descargar Detalle Back Office (Excel)",
                    data=df_to_excel_bytes(df_det_bo, "DetalleBackOffice"),
                    file_name=f"detalle_backoffice_{day_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ==================== TAB 2: CANCELADAS ====================
    with tabs[2]:
        st.subheader("Canceladas (Canc Error)")

        df_canc = df[df["Estatus"] == "Canc Error"]
        if df_canc.empty:
            st.info("No hay registros cancelados para los filtros actuales.")
        else:
            by_day = df_canc.groupby("Fecha", as_index=False).size()
            fig = px.bar(
                by_day,
                x="Fecha",
                y="size",
                title="Canceladas por día",
                labels={"size": "Total Canc Error"},
            )
            st.plotly_chart(fig, width="stretch")

            # =========================
            # ✅ Vista por meses y semanas (Canc Error) + Comparativo día vs día + Fecha 1 vs Fecha 2
            # =========================
            df_canc_ctx = df_canc.copy()
            df_canc_ctx["C_DT"] = pd.to_datetime(df_canc_ctx["Fecha creacion"], errors="coerce")
            df_canc_ctx["C_Fecha"] = df_canc_ctx["C_DT"].dt.date
            df_canc_ctx["C_Hora"] = df_canc_ctx["C_DT"].dt.hour

            df_canc_ctx = df_canc_ctx[df_canc_ctx["C_DT"].notna()].copy()

            if not df_canc_ctx.empty:
                df_canc_ctx["C_MonthKey"] = df_canc_ctx["C_DT"].dt.strftime("%Y-%m")
                df_canc_ctx["C_MonthName"] = df_canc_ctx["C_DT"].dt.strftime("%B")
                df_canc_ctx["C_MonthLabel"] = df_canc_ctx["C_MonthKey"] + " (" + df_canc_ctx["C_MonthName"] + ")"

                month_start = df_canc_ctx["C_DT"].dt.to_period("M").dt.to_timestamp()
                first_wd = month_start.dt.weekday
                df_canc_ctx["C_WeekOfMonth"] = ((df_canc_ctx["C_DT"].dt.day + first_wd - 1) // 7) + 1

                st.markdown("### Vista por meses y semanas (Canc Error)")

                c_month_options = sorted(df_canc_ctx["C_MonthLabel"].dropna().unique().tolist())
                c_default_months = c_month_options if c_month_options else []

                c_months_sel = st.multiselect(
                    "Selecciona uno o más meses (Canc Error)",
                    options=c_month_options,
                    default=c_default_months,
                    key="canc_months_multi",
                )

                df_cmw = df_canc_ctx.copy()
                if c_months_sel:
                    df_cmw = df_cmw[df_cmw["C_MonthLabel"].isin(c_months_sel)].copy()
                else:
                    df_cmw = df_cmw.iloc[0:0].copy()

                if df_cmw.empty:
                    st.info("No hay datos Canc Error para los meses seleccionados.")
                else:
                    df_cmw["C_WeekLabel"] = df_cmw["C_MonthLabel"] + " - Semana " + df_cmw["C_WeekOfMonth"].astype(int).astype(str)
                    c_week_options = sorted(df_cmw["C_WeekLabel"].dropna().unique().tolist())
                    c_default_weeks = c_week_options if c_week_options else []

                    c_weeks_sel = st.multiselect(
                        "Selecciona Semana(s) del mes (Canc Error)",
                        options=c_week_options,
                        default=c_default_weeks,
                        key="canc_weeks_multi",
                    )

                    if c_weeks_sel:
                        df_cmw = df_cmw[df_cmw["C_WeekLabel"].isin(c_weeks_sel)].copy()
                    else:
                        df_cmw = df_cmw.iloc[0:0].copy()

                    if df_cmw.empty:
                        st.info("No hay datos Canc Error para las semanas seleccionadas.")
                    else:
                        by_day_cmw = df_cmw.groupby("C_Fecha", as_index=False).size()
                        fig_cmw = px.bar(
                            by_day_cmw,
                            x="C_Fecha",
                            y="size",
                            title="Total por día (Canc Error) — filtro por Mes(es) y Semana(s)",
                            labels={"size": "Total Canc Error", "C_Fecha": "Fecha"},
                        )
                        st.plotly_chart(fig_cmw, width="stretch")

                        st.markdown("### Comparativo día vs día (mes contra mes) — Canc Error")

                        df_cmw["C_DiaDelMes"] = df_cmw["C_DT"].dt.day
                        cmp_c = (
                            df_cmw.groupby(["C_MonthLabel", "C_DiaDelMes"], as_index=False)
                            .size()
                            .rename(columns={"size": "Total"})
                        )

                        fig_cmp_c = px.line(
                            cmp_c,
                            x="C_DiaDelMes",
                            y="Total",
                            color="C_MonthLabel",
                            markers=True,
                            title="Comparativo por día del mes (Canc Error)",
                            labels={
                                "C_DiaDelMes": "Día del mes",
                                "Total": "Total Canc Error",
                                "C_MonthLabel": "Mes",
                            },
                        )
                        fig_cmp_c.update_xaxes(dtick=1)
                        st.plotly_chart(fig_cmp_c, width="stretch")

                        st.markdown("#### Comparar dos fechas específicas (calendario) — Canc Error")

                        c_avail_dates = sorted(df_cmw["C_Fecha"].dropna().unique().tolist())
                        if not c_avail_dates:
                            st.info("No hay fechas disponibles para comparar con los filtros actuales (Canc Error).")
                        else:
                            c_d_min = c_avail_dates[0]
                            c_d_max = c_avail_dates[-1]
                            c_d_def_2 = c_d_max
                            c_d_def_1 = c_avail_dates[-2] if len(c_avail_dates) >= 2 else c_d_max

                            cA, cB = st.columns(2)
                            with cA:
                                c_d1 = st.date_input(
                                    "Fecha 1 (día a comparar) — Canc Error",
                                    value=c_d_def_1,
                                    min_value=c_d_min,
                                    max_value=c_d_max,
                                    key="canc_cmp_calendar_date1",
                                )
                            with cB:
                                c_d2 = st.date_input(
                                    "Fecha 2 (día a comparar) — Canc Error",
                                    value=c_d_def_2,
                                    min_value=c_d_min,
                                    max_value=c_d_max,
                                    key="canc_cmp_calendar_date2",
                                )

                            df_cd1 = df_cmw[df_cmw["C_Fecha"] == c_d1].copy()
                            df_cd2 = df_cmw[df_cmw["C_Fecha"] == c_d2].copy()

                            ct1 = int(df_cd1.shape[0])
                            ct2 = int(df_cd2.shape[0])

                            cm1, cm2, cm3 = st.columns(3)
                            with cm1:
                                st.metric("Total Canc Error (Fecha 1)", ct1)
                            with cm2:
                                st.metric("Total Canc Error (Fecha 2)", ct2)
                            with cm3:
                                st.metric("Diferencia (Fecha 1 - Fecha 2)", ct1 - ct2)

                            c_comp_df = pd.DataFrame({"Fecha": [str(c_d1), str(c_d2)], "Total": [ct1, ct2]})
                            fig_c_dates = px.bar(
                                c_comp_df,
                                x="Fecha",
                                y="Total",
                                title="Comparativo Canc Error — Fecha vs Fecha",
                                labels={"Total": "Total Canc Error"},
                            )
                            st.plotly_chart(fig_c_dates, width="stretch")

                            ch1 = df_cd1.groupby("C_Hora").size()
                            ch2 = df_cd2.groupby("C_Hora").size()
                            hours = list(range(0, 24))
                            c_hour_df = pd.DataFrame(
                                {
                                    "Hora": hours,
                                    str(c_d1): [int(ch1.get(h, 0)) for h in hours],
                                    str(c_d2): [int(ch2.get(h, 0)) for h in hours],
                                }
                            )
                            c_hour_long = c_hour_df.melt(id_vars="Hora", var_name="Fecha", value_name="Total")

                            fig_c_hour = px.bar(
                                c_hour_long,
                                x="Hora",
                                y="Total",
                                color="Fecha",
                                barmode="group",
                                title="Comparativo por hora — Fecha vs Fecha (Canc Error)",
                                labels={"Total": "Total Canc Error", "Hora": "Hora"},
                            )
                            st.plotly_chart(fig_c_hour, width="stretch")

            day_options = sorted(by_day["Fecha"].unique())
            today = date.today()
            default_index = day_options.index(today) if today in day_options else len(day_options) - 1

            day_sel = st.selectbox(
                "Selecciona un día (Canceladas)",
                day_options,
                index=default_index,
            )
            df_day = df_canc[df_canc["Fecha"] == day_sel]

            by_hour = df_day.groupby("Hora", as_index=False).size()
            fig2 = px.bar(
                by_hour,
                x="Hora",
                y="size",
                title=f"Desglose por hora – {day_sel}",
                labels={"size": "Total Canc Error"},
            )
            st.plotly_chart(fig2, width="stretch")

            by_hour_team = (
                df_day.groupby(["Hora", "Jefe directo"], as_index=False)
                .size()
                .rename(columns={"size": "Total"})
            )

            fig_team_canc = px.bar(
                by_hour_team,
                x="Hora",
                y="Total",
                color="Jefe directo",
                barmode="group",
                title=f"Canceladas por hora y equipo – {day_sel}",
                labels={
                    "Total": "Total Canc Error",
                    "Hora": "Hora",
                    "Jefe directo": "Supervisor",
                },
            )
            st.plotly_chart(fig_team_canc, width="stretch")

            st.subheader("Detalle de cancelaciones (Ejecutivo / Jefe directo)")
            detalle_cols = [
                c
                for c in [
                    "Jefe directo",
                    "Vendedor",
                    "Cliente",
                    "Telefono",
                    "Folio",
                    "Fecha",
                    "Hora",
                    "Centro",
                    "Estatus",
                    "Venta",
                ]
                if c in df_day.columns
            ]

            df_det = df_day[detalle_cols].rename(
                columns={
                    "Vendedor": "Ejecutivo",
                    "Telefono": "Telefono cliente",
                }
            )
            df_det = df_det.sort_values(
                [col for col in ["Jefe directo", "Ejecutivo", "Hora", "Folio"] if col in df_det.columns]
            )

            st.dataframe(df_det, width="stretch")

            st.download_button(
                "Descargar Detalle Canceladas (Excel)",
                data=df_to_excel_bytes(df_det, "DetalleCanceladas"),
                file_name=f"detalle_canceladas_{day_sel}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ==================== TAB 3: PROGRAMADAS X SEMANA ====================
    with tabs[3]:
        st.subheader("Programadas por semana")

        df_prog_base = df_no_month[df_no_month["Estatus"] != "Canc Error"].copy()

        if df_prog_base.empty:
            st.info("No hay programadas para los filtros actuales.")
        else:
            df_prog = df_prog_base

            if mes_sel != "All":
                month_rows = df_no_month[df_no_month["Mes"] == mes_sel].copy()
                if not month_rows.empty:
                    start_ts = pd.to_datetime(month_rows["Fecha creacion"], errors="coerce").min()
                    end_ts = pd.to_datetime(month_rows["Fecha creacion"], errors="coerce").max()

                    if pd.notna(start_ts) and pd.notna(end_ts):
                        week_start = start_ts - pd.Timedelta(days=int(start_ts.weekday()))
                        week_end = end_ts + pd.Timedelta(days=int(6 - end_ts.weekday()))

                        df_prog = df_prog_base[
                            (df_prog_base["Fecha creacion"] >= week_start)
                            & (df_prog_base["Fecha creacion"] <= week_end)
                        ].copy()

            by_week = df_prog.groupby("Año Semana", as_index=False).size()
            fig = px.bar(
                by_week,
                x="Año Semana",
                y="size",
                title="Vista general de programadas por semana",
                labels={"size": "Total Programadas"},
            )
            fig.update_xaxes(type="category")
            st.plotly_chart(fig, width="stretch")

            st.download_button(
                "Descargar Programadas (Excel)",
                data=df_to_excel_bytes(df_prog, "Programadas"),
                file_name=f"programadas_{fecha_ini}_{fecha_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ==================== TAB 4: PROGRAMADAS – TOP EJECUTIVOS ====================
    with tabs[4]:
        st.subheader("Top Ejecutivos – Programadas")

        df_prog = df[df["Estatus"] != "Canc Error"]
        if df_prog.empty:
            st.info("No hay programadas para los filtros actuales.")
        else:
            by_exec_all = (
                df_prog.groupby("Vendedor", as_index=False)
                .size()
                .rename(columns={"size": "Total Programadas"})
                .sort_values("Total Programadas", ascending=False)
            )

            by_exec = by_exec_all.head(30)

            n_exec = len(by_exec)
            row_height = 26
            fig_height = max(400, n_exec * row_height + 120)

            fig = px.bar(
                by_exec,
                x="Total Programadas",
                y="Vendedor",
                orientation="h",
                title="Top Ejecutivos Global (Programadas)",
                labels={"Vendedor": "Ejecutivo"},
            )
            fig.update_layout(
                height=fig_height,
                margin=dict(l=260, r=40, t=60, b=40),
                yaxis=dict(automargin=True),
            )
            st.plotly_chart(fig, width="stretch")

            st.subheader("Ranking completo (todos los ejecutivos)")
            st.dataframe(by_exec_all, width="stretch")

            st.download_button(
                "Descargar Top Ejecutivos (Excel)",
                data=df_to_excel_bytes(by_exec_all, "TopEjecutivos"),
                file_name=f"top_ejecutivos_{fecha_ini}_{fecha_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ==================== TAB 5: DETALLE GENERAL ====================
    with tabs[5]:
        st.subheader("Detalle general programadas")

        if df.empty:
            st.info("Sin datos para los filtros actuales.")
        else:
            df_flags = df.copy()

            df_flags["flag_Programada"] = (df_flags["Estatus"] != "Canc Error").astype(int)
            df_flags["flag_Activadas"] = (df_flags["Status"] == "Entregado").astype(int)
            df_flags["flag_EnTransito"] = (df_flags["Status"] == "En Transito").astype(int)
            df_flags["flag_ET_EnEntrega"] = (
                (df_flags["Status"] == "En Transito") & (df_flags["Estatus"] == "En entrega")
            ).astype(int)
            df_flags["flag_ET_EnPreparacion"] = (
                (df_flags["Status"] == "En Transito") & (df_flags["Estatus"] == "En preparacion")
            ).astype(int)
            df_flags["flag_ET_Solicitado"] = (
                (df_flags["Status"] == "En Transito") & (df_flags["Estatus"] == "Solicitado")
            ).astype(int)
            df_flags["flag_ET_BackOffice"] = (
                (df_flags["Status"] == "En Transito") & (df_flags["Estatus"] == "Back Office")
            ).astype(int)
            df_flags["flag_ET_EntregadoSinVenta"] = (
                (df_flags["Status"] == "En Transito") & (df_flags["Estatus"] == "Entregado")
            ).astype(int)

            agg_dict = {
                "TotalProgramadas": ("flag_Programada", "sum"),
                "Activadas": ("flag_Activadas", "sum"),
                "EnTransito": ("flag_EnTransito", "sum"),
                "ET En entrega": ("flag_ET_EnEntrega", "sum"),
                "ET En preparacion": ("flag_ET_EnPreparacion", "sum"),
                "ET Solicitado": ("flag_ET_Solicitado", "sum"),
                "ET Back Office": ("flag_ET_BackOffice", "sum"),
                "ET Entregado sin venta": ("flag_ET_EntregadoSinVenta", "sum"),
            }

            grouped = (
                df_flags.groupby(["Jefe directo", "Vendedor"], as_index=False)
                .agg(**agg_dict)
                .rename(columns={"Vendedor": "Ejecutivo"})
            )
            grouped = grouped.sort_values(["Jefe directo", "Ejecutivo"])

            metric_cols = [
                "TotalProgramadas",
                "Activadas",
                "EnTransito",
                "ET En entrega",
                "ET En preparacion",
                "ET Solicitado",
                "ET Back Office",
                "ET Entregado sin venta",
            ]

            total_row = {"Jefe directo": "Total", "Ejecutivo": ""}
            for col in metric_cols:
                total_row[col] = int(grouped[col].sum())

            grouped_with_total = pd.concat([grouped, pd.DataFrame([total_row])], ignore_index=True)

            if "EnTransito" in grouped_with_total.columns:
                styled_grouped = grouped_with_total.style.set_properties(
                    subset=["EnTransito"],
                    **{
                        "background-color": "rgba(34,197,94,0.22)",
                        "font-weight": "800",
                    },
                )
                st.dataframe(styled_grouped, width="stretch")
            else:
                st.dataframe(grouped_with_total, width="stretch")

            by_sup = (
                grouped.groupby("Jefe directo", as_index=False)["TotalProgramadas"]
                .sum()
                .rename(columns={"Jefe directo": "Supervisor"})
            )
            fig = px.pie(
                by_sup,
                names="Supervisor",
                values="TotalProgramadas",
                title="Programadas por supervisor",
            )
            fig.update_traces(textposition="inside", textinfo="label+percent")
            fig.update_layout(showlegend=True)
            st.plotly_chart(fig, width="stretch")

            st.download_button(
                "Descargar detalle general (Excel)",
                data=df_to_excel_bytes(grouped_with_total, "DetalleGeneral"),
                file_name=f"detalle_general_{fecha_ini}_{fecha_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.subheader("Detalle de registros En Tránsito")

            cols_det_en_t = [
                c
                for c in [
                    "Jefe directo",
                    "Vendedor",
                    "Folio",
                    "Fecha",
                    "Hora",
                    "Centro",
                    "Estatus",
                    "Status",
                    "Venta",
                ]
                if c in df.columns
            ]

            df_en_t = df[df["Status"] == "En Transito"][cols_det_en_t].rename(
                columns={"Vendedor": "Ejecutivo"}
            )

            if df_en_t.empty:
                st.info("No hay registros En Transito para los filtros actuales.")
            else:
                df_en_t = df_en_t.sort_values(
                    [col for col in ["Jefe directo", "Ejecutivo", "Fecha", "Hora", "Folio"] if col in df_en_t.columns]
                )
                st.dataframe(df_en_t, width="stretch")

                st.download_button(
                    "Descargar detalle En Tránsito (Excel)",
                    data=df_to_excel_bytes(df_en_t, "EnTransitoDetalle"),
                    file_name=f"en_transito_detalle_{fecha_ini}_{fecha_fin}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ==================== TAB 6: SIN VENTA ====================
    with tabs[6]:
        st.subheader("Ejecutivos sin venta (mes actual)")

        total_sinv = kpi_total_sinventa(sinv_fil)
        st.metric("Total ejecutivos sin venta", total_sinv)

        if sinv_fil.empty:
            st.info("No hay ejecutivos sin venta para los filtros actuales.")
        else:
            df_sinv = sinv_fil.sort_values(["JefeDirecto", "NombreCompleto"])
            st.dataframe(df_sinv[["JefeDirecto", "NombreCompleto"]], width="stretch")

            st.download_button(
                "Descargar Sin Venta (Excel)",
                data=df_to_excel_bytes(df_sinv[["JefeDirecto", "NombreCompleto"]], "SinVenta"),
                file_name=f"sin_venta_{fecha_ini}_{fecha_fin}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

if __name__ == "__main__":
    main()
